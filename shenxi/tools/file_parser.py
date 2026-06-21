import os


def parse_file(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".txt":
        return _read_text(file_path)
    elif ext == ".pdf":
        return _read_pdf(file_path)
    elif ext == ".docx":
        return _read_docx(file_path)
    elif ext in (".xls", ".xlsx"):
        return _read_excel(file_path)
    return f"不支持的文件格式：{ext}"


def _read_text(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


def _read_pdf(path: str) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(path)
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    except ImportError:
        return "（PDF 解析组件未安装）"
    except Exception as e:
        return f"（PDF 解析失败：{e}）"


def _read_docx(path: str) -> str:
    try:
        from docx import Document
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        return "（Word 解析组件未安装）"
    except Exception as e:
        return f"（Word 解析失败：{e}）"


def _read_excel(path: str) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"【{sheet_name}】")
            for row in ws.iter_rows(values_only=True):
                parts.append("\t".join(str(c or "") for c in row))
        return "\n".join(parts)
    except ImportError:
        return "（Excel 解析组件未安装）"
    except Exception as e:
        return f"（Excel 解析失败：{e}）"
